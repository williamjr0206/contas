"""
Importa a planilha PREVISAO.xlsx para o banco MySQL/MariaDB szjw_contas.

Tabelas usadas:
- grupos(id_grupo, descricao)
- lancamentos(id_lancamento, documento_numero, data_lancamento, descricao, tipo,
              data_vencimento, valor_nominal, data_pagamento, valor_pago,
              status, forma_de_pagamento_recebimento, id_grupo)

Antes de rodar, instale as dependências:
    pip install openpyxl mysql-connector-python

Ajuste abaixo USUARIO, SENHA, HOST e CAMINHO_PLANILHA conforme seu ambiente.
"""

from __future__ import annotations

import csv
import re
import unicodedata
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path

import mysql.connector
from openpyxl import load_workbook

# =========================
# CONFIGURAÇÕES
# =========================
HOST = "localhost"
USUARIO = "root"
SENHA = ""          # No XAMPP geralmente fica vazio. Na Milbr, coloque a senha do usuário MySQL.
BANCO = "szjw_contas"
PORTA = 3306

CAMINHO_PLANILHA = Path("PREVISAO.xlsx")
ABAS_LANCAMENTOS = ["WILLIAM"]  # Zenilda foi desconsiderada conforme solicitado
ABA_GRUPOS = "GRUPOS"

# Se estiver testando, deixe True. Ele mostra o que faria, mas NÃO grava no banco.
MODO_TESTE = False

# Se True, apaga os registros atuais antes de importar.
# Cuidado: use apenas se você quiser recriar a importação do zero.
LIMPAR_TABELAS_ANTES = False

ARQUIVO_LOG = Path("log_importacao_previsao.csv")


# =========================
# FUNÇÕES AUXILIARES
# =========================
def normalizar_texto(valor) -> str:
    if valor is None:
        return ""
    texto = str(valor).strip()
    texto = re.sub(r"\s+", " ", texto)
    return texto


def chave_normalizada(valor) -> str:
    texto = normalizar_texto(valor).upper()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    return texto.strip()


def excel_para_data(valor):
    """Converte data vinda do Excel para date ou None."""
    if valor in (None, ""):
        return None

    if isinstance(valor, datetime):
        return valor.date()

    if isinstance(valor, date):
        return valor

    # Excel serial date: 1 = 1900-01-01, mas Excel tem bug histórico de 1900.
    if isinstance(valor, (int, float)):
        return (datetime(1899, 12, 30) + timedelta(days=float(valor))).date()

    texto = normalizar_texto(valor)
    if not texto:
        return None

    formatos = ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"]
    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"Data inválida: {valor!r}")


def para_decimal(valor):
    if valor in (None, ""):
        return None
    if isinstance(valor, Decimal):
        return valor
    if isinstance(valor, (int, float)):
        return Decimal(str(valor))

    texto = normalizar_texto(valor)
    if not texto:
        return None

    # Aceita formatos como 1.234,56 ou 1234.56
    texto = texto.replace("R$", "").strip()
    if "," in texto:
        texto = texto.replace(".", "").replace(",", ".")

    try:
        return Decimal(texto)
    except InvalidOperation:
        raise ValueError(f"Valor monetário inválido: {valor!r}")


def mapear_tipo(valor) -> str:
    mapa = {
        "PAGAR": "Pagar",
        "RECEBER": "Receber",
    }
    chave = chave_normalizada(valor)
    if chave not in mapa:
        raise ValueError(f"Tipo inválido: {valor!r}")
    return mapa[chave]


def mapear_status(valor) -> str:
    mapa = {
        "ABERTO": "Aberto",
        "PAGA": "Pago",
        "PAGO": "Pago",
        "RECEBIDA": "Recebido",
        "RECEBIDO": "Recebido",
    }
    chave = chave_normalizada(valor)
    if chave not in mapa:
        raise ValueError(f"Status inválido: {valor!r}")
    return mapa[chave]


def mapear_forma_pagamento(valor, tipo: str):
    """
    A tabela aceita somente estes ENUMs:
    Pix Recebido, Pix QR Code, Aplicação, Cartão Débito, Débito Automático,
    Crédito em Conta, Débito em Conta, Pagamento Boleto, Pix Pagamento,
    Transação Bancária.

    Formas que não existem no ENUM, como CARTÃO CRÉDITO e DINHEIRO,
    ficam como NULL e aparecem no log para você decidir depois se altera o ENUM.
    """
    if valor in (None, ""):
        return None

    original = normalizar_texto(valor)
    chave = chave_normalizada(original)

    mapa_fixo = {
        "PIX QR CODE": "Pix QR Code",
        "APLICACAO": "Aplicação",
        "CARTAO DE DEBITO": "Cartão Débito",
        "CARTAO DEBITO": "Cartão Débito",
        "DEBITO AUTOMATICO": "Débito Automático",
        "DEBITO EM CONTA": "Débito em Conta",
        "PAGAMENTO BOLETO": "Pagamento Boleto",
        "PAGAMENTO DE BOLETO": "Pagamento Boleto",
        "TRANSACAO BANCARIA": "Transação Bancária",
        "MOVIMENTACAO FINANCEIRA": "Transação Bancária",
        "MOVIMENTACAO FINANCEIRA": "Transação Bancária",
        "DEPOSITO EM CONTA": "Crédito em Conta",
        "CREDITO EM CONTA": "Crédito em Conta",
        "SALDO EM CONTA": "Crédito em Conta",
        "ENTRADA": "Crédito em Conta",
        "RESGATE": "Aplicação",
        "RENDIMENTOS": "Aplicação",
        "PAGAMENTO DE CAPITALIZACAO": "Aplicação",
        "PIX RECEBIDO": "Pix Recebido",
        "PIX ENVIADO": "Pix Pagamento",
    }

    if chave == "PIX":
        return "Pix Recebido" if tipo == "Receber" else "Pix Pagamento"

    return mapa_fixo.get(chave)


def linha_vazia(ws, linha: int) -> bool:
    return all(ws.cell(linha, col).value in (None, "") for col in range(1, 12))


# =========================
# LEITURA DA PLANILHA
# =========================
def ler_grupos(wb):
    ws = wb[ABA_GRUPOS]
    grupos = []
    mapa_descricao_para_id = {}

    for linha in range(2, ws.max_row + 1):
        id_grupo = ws.cell(linha, 1).value
        descricao = normalizar_texto(ws.cell(linha, 2).value)

        if id_grupo in (None, "") and not descricao:
            continue
        if id_grupo in (None, "") or not descricao:
            print(f"[AVISO] Grupo ignorado na linha {linha}: id ou descrição vazio.")
            continue

        id_grupo = int(id_grupo)
        grupos.append((id_grupo, descricao))
        mapa_descricao_para_id[chave_normalizada(descricao)] = id_grupo

    return grupos, mapa_descricao_para_id


def ler_lancamentos(wb, mapa_grupos):
    lancamentos = []
    avisos = []

    for aba in ABAS_LANCAMENTOS:
        ws = wb[aba]

        for linha in range(2, ws.max_row + 1):
            if linha_vazia(ws, linha):
                continue

            try:
                documento_numero = normalizar_texto(ws.cell(linha, 1).value)
                data_lancamento = excel_para_data(ws.cell(linha, 2).value)
                descricao = normalizar_texto(ws.cell(linha, 3).value)[:100]
                valor_nominal = para_decimal(ws.cell(linha, 4).value)
                data_vencimento = excel_para_data(ws.cell(linha, 5).value)
                tipo = mapear_tipo(ws.cell(linha, 6).value)
                data_pagamento = excel_para_data(ws.cell(linha, 7).value)
                valor_pago = para_decimal(ws.cell(linha, 8).value)
                status = mapear_status(ws.cell(linha, 9).value)
                forma_original = normalizar_texto(ws.cell(linha, 10).value)
                forma = mapear_forma_pagamento(forma_original, tipo)
                grupo_descricao = normalizar_texto(ws.cell(linha, 11).value)
                id_grupo = mapa_grupos.get(chave_normalizada(grupo_descricao))

                if not documento_numero:
                    documento_numero = f"{aba}-{linha}"

                if not data_lancamento:
                    raise ValueError("data_lancamento vazia")
                if not data_vencimento:
                    raise ValueError("data_vencimento vazia")
                if not descricao:
                    raise ValueError("descrição vazia")
                if valor_nominal is None:
                    raise ValueError("valor nominal vazio")
                if id_grupo is None:
                    raise ValueError(f"grupo não encontrado: {grupo_descricao!r}")

                if forma_original and forma is None:
                    avisos.append({
                        "aba": aba,
                        "linha": linha,
                        "campo": "forma_de_pagamento_recebimento",
                        "valor_original": forma_original,
                        "acao": "Importado como NULL porque não existe no ENUM da tabela.",
                    })

                lancamentos.append({
                    "aba_origem": aba,
                    "linha_origem": linha,
                    "documento_numero": documento_numero[:15],
                    "data_lancamento": data_lancamento,
                    "descricao": descricao,
                    "tipo": tipo,
                    "data_vencimento": data_vencimento,
                    "valor_nominal": valor_nominal,
                    "data_pagamento": data_pagamento,
                    "valor_pago": valor_pago,
                    "status": status,
                    "forma_de_pagamento_recebimento": forma,
                    "id_grupo": id_grupo,
                })

            except Exception as e:
                avisos.append({
                    "aba": aba,
                    "linha": linha,
                    "campo": "linha inteira",
                    "valor_original": "",
                    "acao": f"NÃO IMPORTADA: {e}",
                })

    return lancamentos, avisos


# =========================
# BANCO DE DADOS
# =========================
def conectar():
    return mysql.connector.connect(
        host=HOST,
        user=USUARIO,
        password=SENHA,
        database=BANCO,
        port=PORTA,
        charset="utf8",
        use_unicode=True,
    )


def gravar_log(avisos):
    with ARQUIVO_LOG.open("w", newline="", encoding="utf-8-sig") as f:
        campos = ["aba", "linha", "campo", "valor_original", "acao"]
        writer = csv.DictWriter(f, fieldnames=campos, delimiter=";")
        writer.writeheader()
        writer.writerows(avisos)


def importar_no_banco(grupos, lancamentos):
    conexao = conectar()
    cursor = conexao.cursor()

    try:
        if LIMPAR_TABELAS_ANTES:
            cursor.execute("DELETE FROM lancamentos")
            cursor.execute("DELETE FROM grupos")

        # Grupos: insere ou atualiza descrição se o id já existir.
        sql_grupos = """
            INSERT INTO grupos (id_grupo, descricao)
            VALUES (%s, %s)
            ON DUPLICATE KEY UPDATE descricao = VALUES(descricao)
        """
        cursor.executemany(sql_grupos, grupos)

        cursor.execute("SELECT COALESCE(MAX(id_lancamento), 0) FROM lancamentos")
        proximo_id = int(cursor.fetchone()[0]) + 1

        sql_lancamentos = """
            INSERT INTO lancamentos
            (id_lancamento, documento_numero, data_lancamento, descricao, tipo,
             data_vencimento, valor_nominal, data_pagamento, valor_pago, status,
             forma_de_pagamento_recebimento, id_grupo)
            VALUES
            (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """

        dados = []
        for item in lancamentos:
            dados.append((
                proximo_id,
                item["documento_numero"],
                item["data_lancamento"],
                item["descricao"],
                item["tipo"],
                item["data_vencimento"],
                float(item["valor_nominal"]),
                item["data_pagamento"],
                float(item["valor_pago"]) if item["valor_pago"] is not None else None,
                item["status"],
                item["forma_de_pagamento_recebimento"],
                item["id_grupo"],
            ))
            proximo_id += 1

        cursor.executemany(sql_lancamentos, dados)
        conexao.commit()

    except Exception:
        conexao.rollback()
        raise
    finally:
        cursor.close()
        conexao.close()


# =========================
# EXECUÇÃO
# =========================
def main():
    if not CAMINHO_PLANILHA.exists():
        raise FileNotFoundError(f"Planilha não encontrada: {CAMINHO_PLANILHA.resolve()}")

    wb = load_workbook(CAMINHO_PLANILHA, data_only=True)

    for aba in [ABA_GRUPOS, *ABAS_LANCAMENTOS]:
        if aba not in wb.sheetnames:
            raise ValueError(f"Aba não encontrada na planilha: {aba}")

    grupos, mapa_grupos = ler_grupos(wb)
    lancamentos, avisos = ler_lancamentos(wb, mapa_grupos)
    gravar_log(avisos)

    print("Resumo da importação")
    print("---------------------")
    print(f"Grupos encontrados: {len(grupos)}")
    print(f"Lançamentos válidos: {len(lancamentos)}")
    print(f"Avisos/linhas não importadas: {len(avisos)}")
    print(f"Log: {ARQUIVO_LOG.resolve()}")

    if MODO_TESTE:
        print("\nMODO_TESTE=True: nada foi gravado no banco.")
        print("Depois de conferir o resumo e o log, mude MODO_TESTE para False e rode novamente.")
        return

    importar_no_banco(grupos, lancamentos)
    print("\nImportação concluída com sucesso!")


if __name__ == "__main__":
    main()
